"""
FALCON2 - チャット表示・テーブルエクスポート Mixin
MainWindow から分離した Mixin クラス。
MainWindow のみが継承することを前提とし、self.* 属性は MainWindow.__init__ で初期化される。
"""
from __future__ import annotations

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import TYPE_CHECKING, Optional, Literal, List, Dict, Any, Tuple, Union
from pathlib import Path
from datetime import datetime, timedelta, date
import json
import logging
import re
import io
import threading
import webbrowser
import tempfile
import html

from settings_manager import SettingsManager

try:
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    import numpy as np
    NUMPY_AVAILABLE = True
except ImportError:
    NUMPY_AVAILABLE = False

if TYPE_CHECKING:
    from db.db_handler import DBHandler
    from modules.formula_engine import FormulaEngine
    from modules.rule_engine import RuleEngine

logger = logging.getLogger(__name__)

class ChatViewMixin:
    """Mixin: FALCON2 - チャット表示・テーブルエクスポート Mixin"""
    def add_message(self, role: Literal["user", "ai", "system"], text: str):
        """
        メッセージを追加（ChatGPT風UI）
        
        Args:
            role: メッセージの役割 ("user", "ai", "system")
            text: メッセージテキスト
        """
        # messages_frame が存在する場合のみ追加
        if not hasattr(self, 'chat_messages_frame') or self.chat_messages_frame is None:
            print(f"[{role}] {text}")
            return
        
        try:
            # メッセージカードを作成
            message_card = self._create_message_card(role, text)
            
            # messages_frameに追加
            message_card.pack(fill=tk.X, pady=8)
            
            # Canvasのスクロール領域を更新
            self.chat_messages_frame.update_idletasks()
            self.chat_canvas.config(scrollregion=self.chat_canvas.bbox("all"))
            
            # 最下部にスクロール
            self.chat_canvas.yview_moveto(1.0)
            
        except Exception as e:
            logging.error(f"メッセージ追加エラー: {e}")
            import traceback
            traceback.print_exc()
        
        print(f"[{role}] {text}")
    
    def _is_table_or_numeric_content(self, text: str) -> bool:
        """
        テキストが表形式または数値が多いかを判定
        
        Args:
            text: メッセージテキスト
        
        Returns:
            表形式または数値が多い場合はTrue
        """
        lines = text.split('\n')
        if len(lines) < 2:
            return False
        
        # CSV形式（カンマ区切り）を検出
        comma_count = sum(line.count(',') for line in lines[:5])  # 最初の5行をチェック
        if comma_count >= 3:
            return True
        
        # 数値が多い行を検出（複数行にわたって数値が多く含まれる場合）
        numeric_line_count = 0
        for line in lines[:10]:  # 最初の10行をチェック
            # 数字、カンマ、スペース、タブが多く含まれる行
            numeric_chars = sum(1 for c in line if c.isdigit() or c in ', \t')
            if len(line) > 0 and numeric_chars / len(line) > 0.4:
                numeric_line_count += 1
        
        if numeric_line_count >= 3:
            return True
        
        return False
    
    def _create_message_card(self, role: Literal["user", "ai", "system"], text: str) -> tk.Frame:
        """
        メッセージカード（Frame + Label）を作成
        
        Args:
            role: メッセージの役割
            text: メッセージテキスト
        
        Returns:
            メッセージカードのFrame
        """
        # メッセージカードのFrame
        card_frame = tk.Frame(self.chat_messages_frame, bg="#FFFFFF")
        
        # フォントを決定（数値・表の場合は等幅フォント）
        if self._is_table_or_numeric_content(text):
            font_family = "Courier New"
        else:
            font_family = "Segoe UI"
        
        # ユーザーメッセージ（右寄せ）
        if role == "user":
            # 右寄せ用のコンテナ
            container = tk.Frame(card_frame, bg="#FFFFFF")
            container.pack(fill=tk.X, anchor=tk.E, padx=10)
            
            # メッセージラベル（右寄せ、背景色#F5F7FA）
            message_label = tk.Label(
                container,
                text=text,
                font=(font_family, 11),
                bg="#F5F7FA",
                fg="#000000",
                wraplength=500,  # 最大幅を設定
                justify=tk.LEFT,
                anchor=tk.W,
                padx=14,
                pady=12,
                relief=tk.FLAT,
                bd=0
            )
            message_label.pack(side=tk.RIGHT)
        
        # AIメッセージまたはシステムメッセージ（左寄せ）
        else:
            # 左寄せ用のコンテナ
            container = tk.Frame(card_frame, bg="#FFFFFF")
            container.pack(fill=tk.X, anchor=tk.W, padx=10)
            
            # 背景色を決定（システムメッセージは薄いグレー）
            bg_color = "#FAFAFA" if role == "system" else "#FFFFFF"
            
            # メッセージラベル（左寄せ、背景色#FFFFFF or #FAFAFA）
            message_label = tk.Label(
                container,
                text=text,
                font=(font_family, 11),
                bg=bg_color,
                fg="#000000",
                wraplength=500,  # 最大幅を設定
                justify=tk.LEFT,
                anchor=tk.W,
                padx=14,
                pady=12,
                relief=tk.FLAT,
                bd=0
            )
            message_label.pack(side=tk.LEFT)
        
        return card_frame
    
    def _add_chat_message(self, sender: str, message: str, color: Optional[str] = None):
        """
        チャットメッセージを追加（後方互換性のため残す）
        
        Args:
            sender: 送信者名 ("ユーザー", "AI回答", "システム" など)
            message: メッセージ
            color: テキストの色（#RRGGBB形式、Noneの場合はデフォルト）
        """
        # senderをroleに変換
        if sender == "ユーザー":
            role = "user"
        elif sender == "AI回答" or sender == "AI":
            role = "ai"
        else:  # "システム" など
            role = "system"
        
        # 新しいadd_messageメソッドを使用
        self.add_message(role=role, text=message)
    
    def _get_event_display_color(self, event_number: int) -> str:
        """
        イベントの表示色を決定（CowCardと同じロジック）
        
        Args:
            event_number: イベント番号
        
        Returns:
            色（#RRGGBB形式）
        """
        # event_dictionary.json を読み込む
        if not hasattr(self, '_event_dictionary') or self._event_dictionary is None:
            self._event_dictionary = {}
            if self.event_dict_path and self.event_dict_path.exists():
                try:
                    with open(self.event_dict_path, 'r', encoding='utf-8') as f:
                        self._event_dictionary = json.load(f)
                except Exception as e:
                    logging.error(f"event_dictionary.json 読み込みエラー: {e}")
        
        # イベント辞書から情報を取得
        event_str = str(event_number)
        event_dict = self._event_dictionary.get(event_str, {})
        
        # 1. display_color が指定されていればそれを使用（最優先）
        display_color = event_dict.get('display_color')
        if display_color:
            return display_color
        
        # 2. category / outcome に応じてデフォルト色を使用
        category = event_dict.get('category', '')
        
        if category == "CALVING":
            return "#0066cc"  # 青
        elif category == "PREGNANCY":
            outcome = event_dict.get('outcome', '')
            if outcome == "NEGATIVE":
                return "#cc0000"  # 赤
            else:
                return "#008000"  # 緑（POSITIVE または未指定）
        elif category == "REPRODUCTION":
            return "#000000"  # 黒
        
        # 3. それ以外は黒
        return "#000000"
    
    def _display_cow_info_in_chat(self, cow_auto_id: int):
        """
        チャット履歴に個体情報とイベント履歴を表示（色付き）
        
        Args:
            cow_auto_id: 牛の auto_id
        """
        try:
            # 牛の情報を取得
            cow = self.db.get_cow_by_auto_id(cow_auto_id)
            if not cow:
                return
            
            # 個体情報を表示
            cow_id = cow.get('cow_id', '')
            jpn10 = cow.get('jpn10', '')
            brd = cow.get('brd', '')
            lact = cow.get('lact', '')
            rc = cow.get('rc', '')
            rc_formatted = self.format_rc(rc)
            
            self._add_chat_message("システム", f"個体情報: {cow_id} (JPN10: {jpn10}, 品種: {brd}, 産次: {lact}, RC: {rc_formatted})")
            
            # イベント履歴を取得
            events = self.db.get_events_by_cow(cow_auto_id, include_deleted=False)
            
            if not events:
                self._add_chat_message("システム", "イベント履歴: なし")
                return
            
            # イベント名を取得するためのヘルパー関数
            def get_event_name(event_number: int) -> str:
                event_str = str(event_number)
                if hasattr(self, '_event_dictionary') and self._event_dictionary:
                    event_dict = self._event_dictionary.get(event_str, {})
                    return event_dict.get('name_jp', f'イベント{event_number}')
                return f'イベント{event_number}'
            
            # イベント履歴を表示（最新から）
            self._add_chat_message("システム", "イベント履歴:")
            for event in events[:10]:  # 最新10件まで
                event_date = event.get('event_date', '')
                event_number = event.get('event_number')
                note = event.get('note', '')
                
                if event_number is None:
                    continue
                
                event_name = get_event_name(event_number)
                color = self._get_event_display_color(event_number)
                
                # イベント情報を色付きで表示
                display_text = f"  {event_date} {event_name}"
                if note:
                    display_text += f" - {note}"
                
                self._add_chat_message("システム", display_text, color=color)
            
            if len(events) > 10:
                self._add_chat_message("システム", f"  ... 他 {len(events) - 10} 件")
                
        except Exception as e:
            logging.error(f"個体情報表示エラー: {e}")
            import traceback
            traceback.print_exc()
    
    def _get_sorted_table_rows(self):
        """
        現在のソート状態に基づいてソートされた行を取得
        
        Returns:
            ソートされた行のリスト
        """
        # ソートする列と方向を取得
        sort_column = None
        sort_direction = None
        for col, state in self.table_sort_state.items():
            if state is not None:
                sort_column = col
                sort_direction = state
                break
        
        # データをソート
        if sort_column and sort_direction:
            sorted_rows = sorted(
                self.table_original_rows,
                key=lambda row: self._get_sort_key(row.get(sort_column)),
                reverse=(sort_direction == 'desc')
            )
        else:
            # ソートなし（元の順序）
            sorted_rows = self.table_original_rows
        
        return sorted_rows
    
    def _on_export_table_to_excel(self):
        """
        表をエクセルファイルにエクスポート（現在のソート状態を反映）
        """
        if not self.table_original_rows:
            messagebox.showwarning("エクスポート", "エクスポートするデータがありません。")
            return
        
        # ファイル保存ダイアログを表示
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")],
            title="表をエクスポート"
        )
        
        if not file_path:
            return
        
        try:
            # 現在のソート状態に基づいてソートされた行を取得
            sorted_rows = self._get_sorted_table_rows()
            
            # エクセル形式で保存を試みる
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment
                
                # ワークブックを作成
                wb = openpyxl.Workbook()
                ws = wb.active
                
                # コマンド情報を最初の行に追加
                current_row = 1
                if self.current_table_command:
                    ws.cell(row=current_row, column=1, value="コマンド:")
                    ws.cell(row=current_row, column=2, value=self.current_table_command)
                    ws.cell(row=current_row, column=1).font = Font(bold=True)
                    ws.cell(row=current_row, column=2).font = Font(italic=True)
                    current_row += 1
                
                # 期間情報を追加
                if hasattr(self, 'current_table_period') and self.current_table_period:
                    ws.cell(row=current_row, column=1, value="対象期間:")
                    ws.cell(row=current_row, column=2, value=self.current_table_period)
                    ws.cell(row=current_row, column=1).font = Font(bold=True)
                    ws.cell(row=current_row, column=2).font = Font(italic=True)
                    current_row += 1
                
                start_row = current_row + 1 if current_row > 1 else 1
                
                # 列名を取得
                columns = list(self.result_treeview['columns'])
                
                # ヘッダー行を追加
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=start_row, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # データ行を追加（ソートされた順序で）
                for row_idx, row_dict in enumerate(sorted_rows, start=start_row + 1):
                    for col_idx, col_name in enumerate(columns, start=1):
                        value = row_dict.get(col_name, "")
                        # Noneの場合は空文字列に変換
                        if value is None:
                            value = ""
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 列幅を自動調整
                for col_idx, col_name in enumerate(columns, start=1):
                    max_length = len(str(col_name))
                    for row in ws.iter_rows(min_row=start_row + 1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
                
                # ファイルを保存
                wb.save(file_path)
                messagebox.showinfo("エクスポート", f"エクセルファイルに保存しました:\n{file_path}")
                
            except ImportError:
                # openpyxlがインストールされていない場合はCSVとして保存
                import csv
                
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    
                    # コマンド情報を最初の行に追加
                    if self.current_table_command:
                        writer.writerow(["コマンド:", self.current_table_command])
                    
                    # 期間情報を追加
                    if hasattr(self, 'current_table_period') and self.current_table_period:
                        writer.writerow(["対象期間:", self.current_table_period])
                    
                    if self.current_table_command or (hasattr(self, 'current_table_period') and self.current_table_period):
                        writer.writerow([])  # 空行
                    
                    # 列名を取得
                    columns = list(self.result_treeview['columns'])
                    
                    # ヘッダー行を追加
                    writer.writerow(columns)
                    
                    # データ行を追加（ソートされた順序で）
                    for row_dict in sorted_rows:
                        row_data = [row_dict.get(col, "") or "" for col in columns]
                        writer.writerow(row_data)
                
                messagebox.showinfo("エクスポート", f"CSVファイルに保存しました:\n{file_path}\n\n（エクセル形式で保存するにはopenpyxlをインストールしてください）")
                
        except Exception as e:
            logging.error(f"エクスポートエラー: {e}", exc_info=True)
            messagebox.showerror("エクスポートエラー", f"エクスポート中にエラーが発生しました:\n{e}")
    
    def _on_print_table(self):
        """
        表をプリント（HTML形式で出力、現在のソート状態を反映）
        """
        if not self.table_original_rows:
            messagebox.showwarning("プリント", "プリントするデータがありません。")
            return
        
        try:
            # 現在のソート状態に基づいてソートされた行を取得
            sorted_rows = self._get_sorted_table_rows()
            
            # 列名を取得
            columns = list(self.result_treeview['columns'])
            
            # 農場名を取得
            settings = SettingsManager(self.farm_path)
            farm_name = settings.get("farm_name", self.farm_path.name)
            
            # 出力日を取得（YYYY-MM-DD形式）
            output_date = datetime.now().strftime("%Y-%m-%d")
            
            # HTMLテーブルを作成
            html_content = """<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>表をプリント</title>
    <style>
        body {
            font-family: "MS Gothic", "MS PGothic", "Meiryo", sans-serif;
            margin: 20px;
            font-size: 12px;
        }
        h1 {
            font-size: 16px;
            margin-bottom: 10px;
        }
        .farm-info {
            margin-bottom: 10px;
            padding: 8px;
            font-size: 14px;
            font-weight: bold;
        }
        .command-info {
            margin-bottom: 15px;
            padding: 8px;
            background-color: #f0f0f0;
            border-left: 3px solid #333;
        }
        table {
            border-collapse: collapse;
            width: 100%;
            margin-top: 10px;
            page-break-inside: auto;
        }
        thead {
            display: table-header-group;
        }
        tbody {
            display: table-row-group;
        }
        tr {
            page-break-inside: avoid;
            page-break-after: auto;
        }
        th, td {
            border: 1px solid #333;
            padding: 6px 8px;
            text-align: left;
            vertical-align: top;
        }
        th {
            background-color: #e0e0e0;
            font-weight: bold;
            position: sticky;
            top: 0;
            z-index: 10;
        }
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        tr:hover {
            background-color: #f0f0f0;
        }
        @media print {
            body {
                margin: 0;
            }
            .command-info {
                page-break-after: avoid;
            }
            table {
                page-break-inside: auto;
            }
            tr {
                page-break-inside: avoid;
            }
        }
    </style>
</head>
<body>
"""
            
            # 農場名と出力日を追加
            html_content += f'    <div class="farm-info">{html.escape(str(farm_name))}　　　　出力：{output_date}</div>\n'
            
            # コマンド情報を追加
            command_info_text = ""
            if self.current_table_command:
                command_info_text = f"コマンド: {html.escape(str(self.current_table_command))}"
            
            # 期間情報を追加
            if hasattr(self, 'current_table_period') and self.current_table_period:
                if command_info_text:
                    command_info_text += f" | 対象期間: {html.escape(str(self.current_table_period))}"
                else:
                    command_info_text = f"対象期間: {html.escape(str(self.current_table_period))}"
            
            if command_info_text:
                html_content += f'    <div class="command-info">{command_info_text}</div>\n'
            
            # テーブル開始
            html_content += '    <table>\n        <thead>\n            <tr>\n'
            
            # ヘッダー行
            for col in columns:
                html_content += f'                <th>{html.escape(str(col))}</th>\n'
            
            html_content += '            </tr>\n        </thead>\n        <tbody>\n'
            
            # データ行（ソートされた順序で）
            for row_dict in sorted_rows:
                html_content += '            <tr>\n'
                for col in columns:
                    value = row_dict.get(col, "")
                    if value is None:
                        value = ""
                    # HTMLエスケープしてセルに追加
                    html_content += f'                <td>{html.escape(str(value))}</td>\n'
                html_content += '            </tr>\n'
            
            html_content += """        </tbody>
    </table>
</body>
</html>"""
            
            # 一時HTMLファイルを作成
            with tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', suffix='.html', delete=False) as f:
                f.write(html_content)
                temp_file_path = f.name
            
            # ブラウザで開く
            webbrowser.open(f'file://{temp_file_path}')
            
            messagebox.showinfo(
                "プリント",
                "ブラウザで表が開かれました。\n"
                "ブラウザのメニューから「印刷」を選択してプリントしてください。\n"
                "（Ctrl+P または ファイル > 印刷）"
            )
            
        except Exception as e:
            logging.error(f"プリントエラー: {e}", exc_info=True)
            messagebox.showerror("プリントエラー", f"プリント中にエラーが発生しました:\n{e}")

