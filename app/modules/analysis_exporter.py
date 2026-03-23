"""
FALCON2 - AnalysisResultExporter
Phase D - Step 4: 分析モードのSQL実行結果をCSV/Excel出力

【設計原則】
- SQL実行結果をそのまま出力（加工禁止）
- 列名・並び順・数値はSQL結果を完全に保持
- 余計な列追加・変換は禁止
"""

import csv
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl がインストールされていません。Excel出力は使用できません。")


class AnalysisResultExporter:
    """
    分析モードのSQL実行結果をCSV/Excel出力するクラス
    
    Phase D - Step 4: 分析結果の外部出力
    """
    
    def __init__(self):
        """初期化"""
        pass
    
    def export_to_csv(
        self,
        rows: List[Dict[str, Any]],
        columns: List[str],
        filepath: Path
    ) -> tuple[bool, Optional[str]]:
        """
        SQL実行結果をCSVファイルに出力（Phase E-3.1/3.2）
        
        Args:
            rows: SQL実行結果の行データ（0件でもOK）
            columns: 列名リスト
            filepath: 出力ファイルパス
        
        Returns:
            (成功フラグ, エラーメッセージ)
        """
        try:
            # ディレクトリを作成（Phase E-3.2）
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            # CSVファイルに書き出し
            with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                
                # 0件でもヘッダー行のみ出力（Phase E-3.1）
                for row in rows:
                    # 行データをそのまま書き出し（加工なし）
                    writer.writerow(row)
            
            logger.info(f"CSV出力完了: {filepath}")
            return (True, None)
            
        except PermissionError as e:
            error_msg = "ファイルの書き込みに失敗しました。フォルダの権限を確認してください"
            logger.error(f"CSV出力権限エラー: {e}", exc_info=True)
            return (False, error_msg)
        except Exception as e:
            error_msg = "ファイルの書き込みに失敗しました"
            logger.error(f"CSV出力エラー: {e}", exc_info=True)
            return (False, error_msg)
    
    def export_to_excel(
        self,
        rows: List[Dict[str, Any]],
        columns: List[str],
        filepath: Path
    ) -> tuple[bool, Optional[str]]:
        """
        SQL実行結果をExcelファイル（.xlsx）に出力（Phase E-3.1/3.2）
        
        Args:
            rows: SQL実行結果の行データ（0件でもOK）
            columns: 列名リスト
            filepath: 出力ファイルパス
        
        Returns:
            (成功フラグ, エラーメッセージ)
        """
        if not OPENPYXL_AVAILABLE:
            error_msg = "openpyxl がインストールされていません。Excel出力は使用できません。"
            logger.error(error_msg)
            return (False, error_msg)
        
        try:
            # ディレクトリを作成（Phase E-3.2）
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            # ワークブックを作成
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis Result"
            
            # ヘッダー行を書き出し
            for col_idx, column in enumerate(columns, start=1):
                ws.cell(row=1, column=col_idx, value=column)
            
            # データ行を書き出し（0件でもヘッダー行のみ出力、Phase E-3.1）
            for row_idx, row in enumerate(rows, start=2):
                for col_idx, column in enumerate(columns, start=1):
                    value = row.get(column)
                    # 値の型を維持（None、数値、文字列をそのまま）
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # ファイルに保存
            wb.save(filepath)
            
            logger.info(f"Excel出力完了: {filepath}")
            return (True, None)
            
        except PermissionError as e:
            error_msg = "ファイルの書き込みに失敗しました。フォルダの権限を確認してください"
            logger.error(f"Excel出力権限エラー: {e}", exc_info=True)
            return (False, error_msg)
        except Exception as e:
            error_msg = "ファイルの書き込みに失敗しました"
            logger.error(f"Excel出力エラー: {e}", exc_info=True)
            return (False, error_msg)
    
    def generate_filename(
        self,
        template_name: Optional[str],
        start_date: Optional[str],
        end_date: Optional[str],
        extension: str
    ) -> str:
        """
        出力ファイル名を生成
        
        Args:
            template_name: テンプレート名（Noneの場合は"analysis"）
            start_date: 開始日（YYYY-MM-DD形式、None可）
            end_date: 終了日（YYYY-MM-DD形式、None可）
            extension: 拡張子（"csv" または "xlsx"）
        
        Returns:
            ファイル名
        """
        parts = []
        
        # テンプレート名
        if template_name:
            parts.append(template_name)
        else:
            parts.append("analysis")
        
        # 日付範囲
        if start_date and end_date:
            parts.append(start_date)
            parts.append(end_date)
        elif start_date:
            parts.append(start_date)
        elif end_date:
            parts.append(end_date)
        
        # タイムスタンプ（日付がない場合）
        if not start_date and not end_date:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            parts.append(timestamp)
        
        filename = "_".join(parts) + f".{extension}"
        return filename

