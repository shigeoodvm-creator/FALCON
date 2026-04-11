"""
繁殖検診フロー用テンプレート出力

見た目・ルールは「個体一括導入」テンプレート（cow_registration_excel_template）と統一:
 ・データシート1行目: 紺ヘッダー(1F4E79)・白太字・中央
 ・2行目: 記入例（薄青 D9E1F2・斜体・個体一括導入と同じ）
 ・freeze_panes A2（3行目から入力）
 ・「記入方法」シートに説明表
Excel（.xlsx）を推奨。CSV は同じ列名＋記入例の2行のみ。
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Sequence, Tuple, Union

# 個体一括導入テンプレート（build_bulk_cow_registration_template_workbook）と同一
HEADER_FILL = "1F4E79"
SAMPLE_FILL = "D9E1F2"


def _write_csv_utf8_sig(path: Path, rows: List[List[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        for row in rows:
            w.writerow(row)


def _apply_header_and_sample_row(
    ws,
    headers: Sequence[str],
    sample_row: Sequence[Union[str, int, float, None]],
    column_widths: Sequence[float],
) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hf = Font(color="FFFFFF", bold=True, name="Meiryo UI", size=10)
    hfill = PatternFill("solid", fgColor=HEADER_FILL)
    sf = Font(italic=True, color="595959", name="Meiryo UI", size=10)
    sfill = PatternFill("solid", fgColor=SAMPLE_FILL)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sample_align = Alignment(vertical="center", wrap_text=True)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = hf
        cell.alignment = ctr
    ws.row_dimensions[1].height = 22

    for c, v in enumerate(sample_row, 1):
        if c > len(headers):
            break
        cell = ws.cell(row=2, column=c, value=v)
        cell.fill = sfill
        cell.font = sf
        cell.alignment = sample_align

    for i, w in enumerate(column_widths, 1):
        if i <= len(headers):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _add_instruction_sheet(
    wb,
    main_title: str,
    rule_lines: List[str],
    column_descriptions: List[Tuple[str, str, str]],
) -> None:
    """個体一括導入の「記入方法」シートと同じ構成"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    ws2 = wb.create_sheet("記入方法")
    bf = Font(bold=True, name="Meiryo UI")
    nf = Font(name="Meiryo UI", size=10)
    ws2.cell(1, 1, f"FALCON {main_title}").font = Font(bold=True, size=13, name="Meiryo UI")

    r = 3
    for line in rule_lines:
        ws2.cell(r, 1, line).font = bf if line.startswith("■") else nf
        r += 1

    r += 1
    hfill2 = PatternFill("solid", fgColor="D6DCE4")
    for c, v in enumerate(("列名", "必須", "説明"), 1):
        cell = ws2.cell(r, c, v)
        cell.font = bf
        cell.fill = hfill2
    r += 1
    for name, req, desc in column_descriptions:
        ws2.cell(r, 1, name).font = nf
        ws2.cell(r, 2, req).font = nf
        ws2.cell(r, 3, desc).font = nf
        r += 1

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 52


def write_ai_et_template(path: Path) -> None:
    """AI/ET 取り込み用テンプレート（個体一括導入と同一ビジュアル）"""
    suffix = path.suffix.lower()
    headers = [
        "牛のID（4桁・空欄可）",
        "個体識別番号(JPN10)",
        "授精日",
        "種雄牛(SIRE)",
        "授精師コード",
        "胚移植(ET)",
        "授精種類コード",
        "備考",
    ]
    sample = [
        "1234",
        "1234567890",
        datetime.now().strftime("%Y/%m/%d"),
        "EXAMPLE_SIRE",
        "1",
        "",
        "",
        "",
    ]
    widths = [16, 20, 14, 16, 14, 12, 18, 28]

    if suffix not in (".xlsx", ".xls"):
        _write_csv_utf8_sig(path, [list(headers), [str(x) for x in sample]])
        return

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI_ET入力"
    _apply_header_and_sample_row(ws, headers, sample, widths)

    desc = [
        ("牛のID（4桁・空欄可）", "", "管理の4桁。空欄なら右のJPN10の下4桁を使用します。"),
        ("個体識別番号(JPN10)", "◎", "10桁の家畜個体識別番号（数字のみ）"),
        ("授精日", "◎", "YYYY/MM/DD 形式で入力してください。"),
        ("種雄牛(SIRE)", "", "種雄牛名・コード（任意）"),
        ("授精師コード", "", "農場設定の授精師コード（任意）"),
        ("胚移植(ET)", "", "胚移植のとき 1 / はい / ET など。空欄は人工授精(AI)です。"),
        ("授精種類コード", "", "空欄のとき、手入力と同様に繁殖処置「〇日後AI」から自動設定されます。"),
        ("備考", "", "メモ（任意）"),
    ]
    rules = [
        "■ 基本ルール",
        "・3行目以降に1行1イベントで記入してください。",
        "・2行目の記入例は削除してから入力してください。",
        "・1行目のヘッダー行は変更しないでください。",
        "",
        "■ 授精種類について",
        "・授精種類コードを空欄にすると、直近の繁殖処置と農場設定の「〇日後AI」から自動で入ります。",
    ]
    _add_instruction_sheet(wb, "繁殖検診フロー AI/ET 取り込みテンプレート", rules, desc)
    wb.save(path)


def write_calving_template(path: Path) -> None:
    suffix = path.suffix.lower()
    headers = [
        "牛のID（4桁）",
        "個体識別番号(JPN10)",
        "分娩日",
        "分娩の難易度",
        "備考",
        "子牛1品種",
        "子牛1性別",
        "子牛1死産",
        "子牛2品種",
        "子牛2性別",
        "子牛2死産",
        "子牛3品種",
        "子牛3性別",
        "子牛3死産",
    ]
    sample = [
        "1234",
        "1234567890",
        datetime.now().strftime("%Y/%m/%d"),
        "1",
        "",
        "ホルスタイン",
        "F",
        "0",
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    widths = [14, 20, 14, 16, 20, 14, 10, 10, 12, 10, 10, 12, 10, 10]

    if suffix not in (".xlsx", ".xls"):
        _write_csv_utf8_sig(path, [list(headers), [str(x) for x in sample]])
        return

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分娩入力"
    _apply_header_and_sample_row(ws, headers, sample, widths)

    desc = [
        ("牛のID（4桁）", "", "管理の4桁番号"),
        ("個体識別番号(JPN10)", "◎", "10桁の家畜個体識別番号"),
        ("分娩日", "◎", "YYYY/MM/DD 形式"),
        ("分娩の難易度", "", "1〜5 または 自然分娩・介助・難産 など"),
        ("備考", "", "特記事項（任意）"),
        ("子牛1品種", "", "例：ホルスタイン"),
        ("子牛1性別", "", "F=雌 M=雄"),
        ("子牛1死産", "", "1=死産 0=生産"),
        ("子牛2品種・性別・死産", "", "双子のとき"),
        ("子牛3品種・性別・死産", "", "三子のとき"),
    ]
    rules = [
        "■ 基本ルール",
        "・3行目以降に1行1イベントで記入してください。",
        "・2行目の記入例は削除してから入力してください。",
        "・1行目のヘッダー行は変更しないでください。",
        "",
        "■ 補足",
        "・一括取り込みでは、手動画面の子牛JPN10登録ダイアログは表示されません。",
    ]
    _add_instruction_sheet(wb, "繁殖検診フロー 分娩取り込みテンプレート", rules, desc)
    wb.save(path)


def write_introduction_template(path: Path) -> None:
    """
    導入一括取り込み（個体一括導入テンプレートと同じ列順・名称をベースに、
    繁殖検診用に「牛のID」「登録日」「タグ」を追加）
    """
    suffix = path.suffix.lower()
    headers = [
        "個体識別番号(JPN10)",
        "牛のID（4桁・空欄可）",
        "品種",
        "生年月日",
        "産次",
        "最終分娩日",
        "最終授精日",
        "最終授精SIRE",
        "授精回数",
        "妊娠状態",
        "群(PEN)",
        "母牛識別番号",
        "登録日",
        "導入日",
        "タグ",
        "メモ",
    ]
    today = datetime.now().strftime("%Y/%m/%d")
    sample = [
        "1234567890",
        "",
        "ホルスタイン",
        "2020/05/15",
        3,
        "2024/09/01",
        "2024/10/15",
        "ABC123",
        2,
        "妊娠鑑定待ち",
        "Lactating",
        "",
        "",
        today,
        "",
        "記入例（この行を削除してから入力してください）",
    ]
    widths = [22, 18, 14, 14, 6, 14, 14, 16, 8, 16, 14, 18, 14, 14, 12, 30]

    if suffix not in (".xlsx", ".xls"):
        _write_csv_utf8_sig(path, [list(headers), [str(x) for x in sample]])
        return

    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "導入入力"
    _apply_header_and_sample_row(ws, headers, sample, widths)

    dv = DataValidation(
        type="list",
        formula1='"受胎なし,妊娠鑑定待ち,妊娠,乾乳"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = "J3:J2000"

    desc = [
        ("個体識別番号(JPN10)", "◎", "10桁の家畜個体識別番号（数字のみ）"),
        ("牛のID（4桁・空欄可）", "", "空欄ならJPN10の6〜9桁目から4桁を使います。"),
        ("品種", "", "ホルスタイン / ジャージー / その他"),
        ("生年月日", "", "YYYY/MM/DD形式"),
        ("産次", "", "0=未経産、1以上=経産牛"),
        ("最終分娩日", "", "経産牛の場合（YYYY/MM/DD）"),
        ("最終授精日", "", "授精済みの場合"),
        ("最終授精SIRE", "", "種雄牛名・コード（任意）"),
        ("授精回数", "", "累積授精回数（数字）"),
        ("妊娠状態", "", "プルダウンから選択（受胎なし・妊娠鑑定待ち・妊娠・乾乳）"),
        ("群(PEN)", "", "群名またはコード"),
        ("母牛識別番号", "", "母牛のJPN10（任意）"),
        ("登録日", "", "空欄のときは導入日と同じ（繁殖指標の起点）"),
        ("導入日", "◎", "外部導入を行った日（YYYY/MM/DD）"),
        ("タグ", "", "一覧表示用（任意）"),
        ("メモ", "", "備考（任意）"),
    ]
    rules = [
        "■ 基本ルール",
        "・3行目以降に1頭1行で記入してください。",
        "・2行目の記入例は削除してから入力してください。",
        "・1行目のヘッダー行は変更しないでください。",
        "・「個体一括導入」テンプレートと同じ列構成です（繁殖検診用の列を追加）。",
        "",
        "■ 妊娠状態について",
        "・「妊娠」は妊娠中として扱います（個体一括導入と同じ）。",
    ]
    _add_instruction_sheet(wb, "繁殖検診フロー 導入取り込みテンプレート", rules, desc)
    wb.save(path)


def write_exit_template(path: Path) -> None:
    suffix = path.suffix.lower()
    headers = [
        "牛のID（4桁）",
        "個体識別番号(JPN10)",
        "退出日",
        "退出の種別",
        "備考",
    ]
    sample = [
        "1234",
        "1234567890",
        datetime.now().strftime("%Y/%m/%d"),
        "売却",
        "",
    ]
    widths = [16, 20, 14, 22, 30]

    if suffix not in (".xlsx", ".xls"):
        _write_csv_utf8_sig(path, [list(headers), [str(x) for x in sample]])
        return

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "退出入力"
    _apply_header_and_sample_row(ws, headers, sample, widths)

    desc = [
        ("牛のID（4桁）", "", "管理の4桁（個体識別番号と併用可）"),
        ("個体識別番号(JPN10)", "", "10桁（推奨）"),
        ("退出日", "◎", "YYYY/MM/DD形式"),
        ("退出の種別", "◎", "売却 / 死亡 / 205 / 206 など"),
        ("備考", "", "任意"),
    ]
    rules = [
        "■ 基本ルール",
        "・3行目以降に1行1イベントで記入してください。",
        "・2行目の記入例は削除してから入力してください。",
        "・1行目のヘッダー行は変更しないでください。",
    ]
    _add_instruction_sheet(wb, "繁殖検診フロー 退出取り込みテンプレート", rules, desc)
    wb.save(path)
