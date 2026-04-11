"""
個体一括導入用 Excel テンプレート（新規農場作成・データ吸い込みで同一仕様）
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

import tkinter as tk
from tkinter import filedialog, messagebox


def build_bulk_cow_registration_template_workbook():
    """
    新規農場作成ウィザードと同一内容の openpyxl Workbook を構築する。
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "個体情報"

    headers = [
        "個体識別番号(JPN10)", "品種", "生年月日", "産次",
        "最終分娩日", "最終授精日", "最終授精SIRE", "授精回数",
        "妊娠状態", "群(PEN)", "母牛識別番号", "導入日", "メモ"
    ]
    hf = Font(color="FFFFFF", bold=True, name="Meiryo UI", size=10)
    hfill = PatternFill("solid", fgColor="1F4E79")
    sf = Font(italic=True, color="595959", name="Meiryo UI", size=10)
    sfill = PatternFill("solid", fgColor="D9E1F2")
    ctr = Alignment(horizontal="center", vertical="center")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = hf
        cell.alignment = ctr
    ws.row_dimensions[1].height = 22

    sample = [
        "1234567890", "ホルスタイン", "2020/05/15", 3,
        "2024/09/01", "2024/10/15", "ABC123", 2,
        "妊娠鑑定待ち", "Lactating", "",
        datetime.now().strftime("%Y/%m/%d"),
        "記入例（この行を削除してから入力してください）",
    ]
    for c, v in enumerate(sample, 1):
        cell = ws.cell(row=2, column=c, value=v)
        cell.fill = sfill
        cell.font = sf

    dv = DataValidation(
        type="list",
        formula1='"受胎なし,妊娠鑑定待ち,妊娠,乾乳"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = "I3:I2000"

    for i, w in enumerate([22, 14, 14, 6, 14, 14, 16, 8, 16, 14, 18, 14, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("記入方法")
    bf = Font(bold=True, name="Meiryo UI")
    nf = Font(name="Meiryo UI")
    ws2.cell(1, 1, "FALCON 個体一括導入テンプレート 記入方法").font = Font(bold=True, size=13, name="Meiryo UI")
    notes = [
        (3, "■ 基本ルール", True),
        (4, "・3行目以降に1頭1行で記入してください", False),
        (5, "・2行目のサンプル行は記入後に削除してください", False),
        (6, "・1行目のヘッダー行は変更しないでください", False),
        (8, "■ 各列の説明", True),
    ]
    for r, txt, bold in notes:
        ws2.cell(r, 1, txt).font = bf if bold else nf
    hfill2 = PatternFill("solid", fgColor="D6DCE4")
    for c, v in enumerate(("列名", "必須", "説明"), 1):
        cell = ws2.cell(9, c, v)
        cell.font = bf
        cell.fill = hfill2
    desc = [
        ("個体識別番号(JPN10)", "◎", "10桁の家畜個体識別番号（数字のみ）"),
        ("品種", "", "ホルスタイン / ジャージー / その他"),
        ("生年月日", "", "YYYY/MM/DD形式"),
        ("産次", "", "0=未経産・子牛、1以上=経産牛"),
        ("最終分娩日", "", "産次1以上の場合（YYYY/MM/DD形式）"),
        ("最終授精日", "", "授精済みの場合（YYYY/MM/DD形式）"),
        ("最終授精SIRE", "", "種雄牛名・コード（任意）"),
        ("授精回数", "", "累積授精回数（数字）"),
        ("妊娠状態", "", "受胎なし / 妊娠鑑定待ち / 妊娠 / 乾乳（プルダウン）"),
        ("群(PEN)", "", "群名またはコード"),
        ("母牛識別番号", "", "母牛のJPN10（任意）"),
        ("導入日", "", "省略時は取り込み実行日"),
        ("メモ", "", "備考（任意）"),
    ]
    for i, row in enumerate(desc, 10):
        for c, v in enumerate(row, 1):
            ws2.cell(i, c, v).font = nf
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 50

    return wb


def prompt_save_bulk_cow_registration_template(parent: Optional[tk.Misc] = None) -> None:
    """
    保存先を尋ねてテンプレートを書き出す。失敗時は messagebox で通知。
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        messagebox.showerror(
            "エラー",
            "テンプレート出力には openpyxl が必要です。\n"
            "コマンドプロンプトで:  pip install openpyxl",
            parent=parent,
        )
        return

    save_path = filedialog.asksaveasfilename(
        parent=parent,
        title="テンプレートを保存",
        defaultextension=".xlsx",
        filetypes=[("Excelファイル", "*.xlsx")],
        initialfile="個体一括導入テンプレート.xlsx",
    )
    if not save_path:
        return

    try:
        wb = build_bulk_cow_registration_template_workbook()
        wb.save(save_path)
        messagebox.showinfo("完了", f"テンプレートを保存しました:\n{save_path}", parent=parent)
    except Exception as e:
        messagebox.showerror("エラー", f"テンプレートの保存に失敗しました:\n{e}", parent=parent)


def save_bulk_cow_registration_template_to_path(path: Path) -> None:
    """指定パスへテンプレートを保存（テスト・バッチ用）。"""
    wb = build_bulk_cow_registration_template_workbook()
    wb.save(str(path))
